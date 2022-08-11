from json import load
from tkinter.font import BOLD
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

def edm(filename):
    wb = load_workbook(filename)
    sheet = wb['Recipients']
    sheet2 = wb['Link Clicks - Detail']

    # Deleting empty rows
    sheet.delete_rows(1,1)
    sheet.delete_rows(2,1)

    # Deleting invalid Group Names
    count = 0

    for row in sheet['C']:
        count += 1
        if row.value == "Reagan_Toh" or row.value == "Seow_Ying":
            sheet.delete_rows(count,2)

    # Insert Status column
    sheet.insert_cols(5)
    sheet['E1'].value = "Status"

    sheet['E1'].font = Font(bold=True)
    sheet['E1'].fill = PatternFill("solid", fgColor="FFFF00")

    # Clicked but unsubbed
    count = 0

    for row in sheet['K']:
        count += 1
        if row.value == "Y":
            sheet[f'E{count}'].value = "1. Clicked (Unsub)"

    # Clicked
    # Getting unique clicked emails
    clicked = []
    count = 0
    for row in sheet2['B']:
        # Exclude header and blank rows
        count += 1
        if count >= 4:
            clicked.append(row.value)

    clicked = set(clicked)

    # Check if specific emails clicked the eDM
    count = 0

    for row in sheet['B']:
        count += 1
        for item in clicked:
            if item == row.value:
                sheet[f'E{count}'].value = "1. Clicked"

    # Opened eDMs
    count = 0
    for row in sheet['H']:
        count += 1
        if row.value and sheet[f'E{count}'].value == None:
            sheet[f'E{count}'].value = "2. Opened"

    # Soft Bounced and Hard Bounced
    count = 0
    for row in sheet['D']:
        count += 1
        if (row.value == 'Remote Server Error' or row.value == 'Error (C)') and sheet[f'E{count}'].value == None:
            sheet[f'E{count}'].value = "4. Soft Bounced"
        elif row.value == 'Bad Address' and sheet[f'E{count}'].value == None:
            sheet[f'E{count}'].value = "5. Hard Bounced"
        elif row.value == 'Sent' and sheet[f'E{count}'].value == None:
            sheet[f'E{count}'].value = "3. Sent"
        

    # Save workbook
    wb.save(filename=filename)
    wb.close()